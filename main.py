import os
import sys
import numpy as np
import pandas as pd
import seaborn as sns
import cv2
import shutil
import tempfile

from PIL import Image
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QPixmap, QImage
from PyQt6.QtWidgets import (
    QApplication, QStackedWidget, QGraphicsScene, QVBoxLayout, QFileDialog, QDialog, QSizePolicy, QMessageBox,
    QTableWidgetItem
)
from PyQt6.uic import loadUi
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from yolov5.detect import Path, run
from openpyxl import load_workbook


def get_app_dir():
    return getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))


class DetectionThread(QThread):
    detection_finished = pyqtSignal(object, list, dict)  # signal includes diagonals
    progress_updated = pyqtSignal(int)

    def __init__(self, file_path, model_path, yolo_path):
        super().__init__()
        self.file_path = file_path
        self.model_path = model_path
        self.yolo_path = yolo_path

    def run(self):
        p, im, diagonals = run(weights=Path(self.model_path), source=Path(self.file_path), progress_callback=self.update_progress)
        self.detection_finished.emit(p, im, diagonals)  # emit diagonals

    def update_progress(self, progress):
        self.progress_updated.emit(progress)


class Params(QDialog):
    """Окно для изменения параметров замера"""

    def __init__(self):
        super().__init__()
        loadUi(os.path.join(get_app_dir(), 'params_seed.ui'), self)
        self.setWindowTitle("Параметры замера")


class NewRes(QDialog):
    """Окно для внесения данных для нового замера"""

    def __init__(self, parent=None):
        super().__init__(parent)
        loadUi(os.path.join(get_app_dir(), 'new_res.ui'), self)
        self.setWindowTitle("Новый замер")
        self.apply_btn.clicked.connect(self.apply_data)

    def apply_data(self):
        try:
            fio = self.fio_input.text()
            date = self.date_input.text()
            time = self.time_input.text()
            supplier = self.supplier_input.text()

            if self.parent() is not None:
                self.parent().set_new_res_data(fio, date, time, supplier)
            self.close()
        except Exception as e:
            print(f"Произошла ошибка в apply_data: {e}")


class MainWindow(QDialog):
    """Класс для взаимодействия с главным окном"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("ROR-tool_seed-1.0")
        loadUi(os.path.join(get_app_dir(), 'main_form_lay_seed.ui'), self)
        self.choose_file.clicked.connect(self.open_file_dialog)
        self.new_res.clicked.connect(self.open_new_res_dialog)
        self.progress_bar.setVisible(False)
        self.start_btn.clicked.connect(self.run_detection)
        self.save_post.clicked.connect(self.save_post_img)
        self.save_check.clicked.connect(self.save_temp_excel)
        self.detection_thread = None
        self.excel_path = None
        self.df = None
        self.temp_path = None

    def closeEvent(self, event):
        """Метод вызывается при закрытии окна."""
        try:
            if self.temp_dir and os.path.exists(self.temp_dir):
                shutil.rmtree(self.temp_dir)
        except Exception as e:
            print(f"Ошибка при удалении временной папки: {e}")
        finally:
            event.accept()

    def save_temp_excel(self):
        # Получаем путь для сохранения измененного файла
        save_path, _ = QFileDialog.getSaveFileName(self, "Сохранить чек-лист", "", "Excel Files (*.xlsx)")

        if save_path:
            try:
                # Убедитесь, что директория существует
                save_dir = os.path.dirname(save_path)
                if not os.path.exists(save_dir):
                    os.makedirs(save_dir)

                # Копирование временного файла в указанный путь
                shutil.copy(self.temp_path, save_path)
            except Exception as e:
                QMessageBox.warning(self, "Ошибка!", f"Ошибка в сохранении файла: {e}")

    def resizeEvent(self, event):
        """Обрабатывает событие изменения размера для подгонки изображений и графика в QGraphicsView."""
        self.fit_images()
        super().resizeEvent(event)

    def fit_images(self):
        """Подгоняет изображения и график к размеру QGraphicsView."""
        if hasattr(self, 'orig_scene') and self.orig_scene:
            self.orig.fitInView(self.orig_scene.sceneRect(), Qt.AspectRatioMode.KeepAspectRatio)
        if hasattr(self, 'post_scene') and self.post_scene:
            self.post.fitInView(self.post_scene.sceneRect(), Qt.AspectRatioMode.KeepAspectRatio)
        if hasattr(self, 'canvas') and self.canvas:
            self.canvas.draw()
            self.canvas.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
            self.canvas.updateGeometry()
            if hasattr(self, 'graph_scene') and self.graph_scene:
                self.plot.fitInView(self.graph_scene.sceneRect(), Qt.AspectRatioMode.KeepAspectRatio)

    def save_post_img(self):
        if hasattr(self, 'qim'):
            file_path, _ = QFileDialog.getSaveFileName(self, 'Сохранить изображение', '', 'Images (*.png *.jpg *.jpeg)')
            if file_path:
                self.qim.save(file_path)

    def open_new_res_dialog(self):
        self.new_res_dialog = NewRes(self)
        self.new_res_dialog.show()

    def set_new_res_data(self, fio, date, time, supplier):
        try:
            self.fio.setText("Имя: " + fio)
            self.date.setText("Дата: " + date)
            self.time.setText("Время: " + time)
            self.sup.setText("Поставщик: " + supplier)
        except Exception as e:
            print(f"Ошибка в set_new_res_data: {e}")

    def open_file_dialog(self):
        fname, _ = QFileDialog.getOpenFileName(self, 'Выбрать изображение', '')
        if fname:
            self.file_path = fname
            self.load_image()

    def load_image(self):
        if self.file_path:
            pixmap = QPixmap(self.file_path)
            self.orig_scene = QGraphicsScene()
            self.orig_scene.addPixmap(pixmap)
            self.orig.setScene(self.orig_scene)
            self.fit_images()

    def run_detection(self):
        if not hasattr(self, 'file_path') or not self.file_path:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, загрузите изображение перед запуском обнаружения.")
            return

        # Очистка старых данных перед новым запуском
        self.clear_previous_results()

        model_path = os.path.join(get_app_dir(), "yolov5/models/modelsweight.pt")
        yolo_path = os.path.join(get_app_dir(), 'yolov5')

        # Запуск детектирования в отдельном потоке
        self.detection_thread = DetectionThread(self.file_path, model_path, yolo_path)
        self.detection_thread.detection_finished.connect(self.display_results)
        self.detection_thread.progress_updated.connect(self.update_progress)
        self.progress_bar.setVisible(True)
        self.detection_thread.start()

    def clear_previous_results(self):
        """Очистка старых данных перед новым запуском детектирования."""
        if hasattr(self, 'post_scene') and self.post_scene:
            self.post_scene.clear()
        if hasattr(self, 'canvas') and self.canvas:
            self.graph_layout.removeWidget(self.canvas)
            self.canvas.deleteLater()
            self.canvas = None
            self.graph_layout.deleteLater()  # Полное удаление старого layout
        self.check_list_widget.clearContents()

    def update_progress(self, progress):
        self.progress_bar.setValue(progress)

    def display_results(self, p, im, diagonals):
        self.progress_bar.setVisible(False)

        # Преобразование изображения из BGR в RGB
        rgb_im = cv2.cvtColor(np.array(im), cv2.COLOR_BGR2RGB)

        detect_img = Image.fromarray(np.array(rgb_im), "RGB")
        qim = QImage(detect_img.tobytes(), detect_img.size[0], detect_img.size[1], QImage.Format.Format_RGB888)
        self.qim = qim
        pixmap = QPixmap.fromImage(qim)
        self.post_scene = QGraphicsScene()
        self.post_scene.addPixmap(pixmap)
        self.post.setScene(self.post_scene)
        self.fit_images()

        # Загрузка и отображение графика распределения размеров
        self.load_and_display_graph(diagonals)

        # Загрузка данных из Excel в таблицу
        self.load_excel_to_table(diagonals)

    def load_and_display_graph(self, diagonals):
        try:
            # Проверяем, есть ли данные для основного зерна в diagonals
            if 'main_grain' in diagonals:
                diams = diagonals['main_grain']
                diams_df = pd.DataFrame(diams, columns=['diameter'])

                # Установка стиля Seaborn
                sns.set(style="whitegrid", palette="muted", color_codes=True)

                if not diams_df.empty:
                    fig = Figure(figsize=(7, 8))  # Увеличение высоты графика
                    ax = fig.add_subplot(111)

                    # Создание гистограммы
                    sns.histplot(data=diams_df, x='diameter', bins=20, kde=False, ax=ax,
                                 color="b", edgecolor="k", alpha=0.7)

                    ax.set_xlabel('Диаметр', fontsize=12)
                    ax.set_ylabel('Частота', fontsize=12)
                    ax.set_title('Распределение диаметров зерен пшеницы', fontsize=12)

                    # Установка размеров шрифтов для меток
                    ax.tick_params(axis='both', which='major', labelsize=10)

                    fig.subplots_adjust(bottom=0.25)  # Увеличение нижнего поля

                    # Удаление старого canvas, если он существует
                    if hasattr(self, 'canvas') and self.canvas:
                        self.graph_layout.removeWidget(self.canvas)
                        self.canvas.deleteLater()
                        self.canvas = None
                        self.graph_layout.deleteLater()  # Полное удаление старого layout

                    # Добавление нового canvas в layout
                    self.canvas = FigureCanvas(fig)
                    self.graph_layout = QVBoxLayout(self.plot)  # Обновление layout для графика
                    self.graph_layout.addWidget(self.canvas)
                    self.plot.setLayout(self.graph_layout)  # Установка layout для графика
                    self.canvas.draw()

                    # Установка минимальных и максимальных размеров для графика
                    self.plot.setMinimumSize(400, 300)
                    self.plot.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
                    self.plot.updateGeometry()

                    self.fit_images()
                else:
                    print("Диаметр dataframe пуст.")
            else:
                print("Нет данных для main_grain в diagonals.")
        except Exception as e:
            print(f"Error in load_and_display_graph: {e}")

    def fill_table(self, diagonals, sheet):
        try:
            if hasattr(self, 'date') and self.date.text():
                sheet['B1'] = self.date.text().split(' ', 1)[1]
            if hasattr(self, 'time') and self.time.text():
                sheet['B2'] = self.time.text().split(' ', 1)[1]
            if hasattr(self, 'fio') and self.fio.text():
                sheet['D1'] = self.fio.text().split(' ', 1)[1]
            if hasattr(self, 'sup') and self.sup.text():
                sheet['D2'] = self.sup.text().split(' ', 1)[1]

            # Расчет и вставка процентного соотношения для каждой линии
            self.calculate_and_insert_areas(diagonals=diagonals, sheet=sheet)

        except IndexError as e:
            print(f"IndexError: {e}. Skipping filling table for missing or invalid data.")
        except Exception as e:
            print(f"Error in fill_table: {e}")

        return sheet

    def calculate_and_insert_areas(self, diagonals, sheet):
        try:
            # Расчет общей площади и для каждого вида отдельно
            total_area = 0.0
            file_areas = {}

            for key, diagonals_list in diagonals.items():
                # Рассчитываем площади для каждого вида
                areas = [(d ** 2) / 2 for d in diagonals_list]
                file_areas[key] = areas
                total_area += sum(areas)

            # Вставка % в ячейки tablewidget
            if total_area > 0:
                for key, areas in file_areas.items():
                    if areas:
                        file_total_area = sum(areas)
                        percentage = (file_total_area / total_area) * 100

                        if key == 'broken_grain':
                            sheet.cell(row=12, column=7).value = f"{percentage:.2f}%"
                        elif key == 'Organic_admixture':
                            sheet.cell(row=13, column=3).value = f"{percentage:.2f}%"
                        elif key == 'Weed_seeds':
                            sheet.cell(row=14, column=3).value = f"{percentage:.2f}%"
                        elif key == 'puny_grain':
                            sheet.cell(row=13, column=7).value = f"{percentage:.2f}%"
                        elif key == 'Barley':
                            sheet.cell(row=17, column=7).value = f"{percentage:.2f}%"
                        elif key == 'Oatmeal':
                             sheet.cell(row=18, column=7).value = f"{percentage:.2f}%"

        except Exception as e:
            print(f"Error in calculate_and_insert_areas: {e}")

    # def calculate_areas(self, file_path):
    #     try:
    #         with open(file_path, 'r') as file:
    #             lines = file.readlines()
    #             areas = [float(line.strip()) ** 2 / 2 for line in lines]
    #         return areas
    #     except Exception as e:
    #         print(f"Error in calculate_areas for {file_path}: {e}")
    #         return []

    def load_excel_to_table(self, diagonals):
        self.excel_path = self.excel_path = os.path.join(get_app_dir(), "check_list.xlsx")

        self.temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(self.temp_dir, "check_list_copy.xlsx")
        self.temp_path = temp_path
        shutil.copy(self.excel_path, temp_path)

        # Загружаем копию книги и выбираем активный лист
        workbook = load_workbook(temp_path)
        sheet = workbook.active

        sheet = self.fill_table(diagonals, sheet)
        workbook.save(self.temp_path)

        # Определяем количество строк и столбцов
        num_rows = sheet.max_row
        num_cols = sheet.max_column

        # Устанавливаем количество строк и столбцов в QTableWidget
        self.check_list_widget.setRowCount(num_rows)
        self.check_list_widget.setColumnCount(num_cols)

        # Устанавливаем заголовки столбцов
        column_headers = []
        for col in range(1, num_cols + 1):
            column_headers.append(str(sheet.cell(row=1, column=col).value))
        self.check_list_widget.setHorizontalHeaderLabels(column_headers)

        # Заполняем таблицу значениями из Excel
        for row in range(1, num_rows + 1):
            for col in range(1, num_cols + 1):
                value = sheet.cell(row=row, column=col).value
                item = QTableWidgetItem(str(value) if value is not None else ' ')
                self.check_list_widget.setItem(row - 1, col - 1, item)



app = QApplication(sys.argv)
main_window = MainWindow()
widget = QStackedWidget()
widget.addWidget(main_window)
widget.show()
widget.move(QApplication.primaryScreen().geometry().center() - widget.rect().center())
try:
    app.exec()
except Exception as e:
    print(f"Error: {e}")
finally:
    print("Exit!")
